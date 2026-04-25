"""
Fyll inn Excel-templaten (Boligkalkulator) med data for én eiendom.

Template-struktur (Bolig-arket):
  C2  = Kjøpesum (prisantydning)
  C4  = Dokumentavgift (andel av kjøpesum, f.eks. 0.025)
  C5  = Transaksjonskostnader (omkostninger i NOK)
  G5  = Egenkapital
  G7  = Oppstartskostnad
  G8  = Ombyggingskostnad
  G9  = Forsikring (NOK/mnd – månedlig, trekkes kun én gang i januar)
  G10 = Eiendomsskatt (NOK/år)
  G11 = Kommunale avgifter (NOK/år)
  G12 = Sameiekostnader / felleskost, pr. måned
  C7  = Rentenivå (desimalbrøk, f.eks. 0.055)
  K2  = Adresse / tittel
  K6  = TV og nett, pr. mnd
  K8  = Strøm i kr. pr. mnd
  K9  = Energiklasse
  C13 = Månedsleie (estimert)
  C15 = Vedlikeholdskostnad (andel av leie, f.eks. 0.05)
"""
from __future__ import annotations

import io
import math
import shutil
import tempfile
from pathlib import Path

from eiendom_analyse_claude.models import RealEstate
from eiendom_analyse_claude.analysis.cashflow_params import CashflowParams


def _safe(val, default=0.0) -> float:
    """Returner float, eller default ved NaN/None."""
    try:
        v = float(val)
        return default if math.isnan(v) else v
    except Exception:
        return default


def fill_boligkalkulator(
    estate: RealEstate,
    rent_monthly: float,
    params: CashflowParams,
    template_path: str,
    tv_net_monthly: float = 800.0,
) -> bytes:
    """
    Kopier Excel-templaten og fyll inn data for én eiendom.

    Parametere
    ----------
    estate : RealEstate
        Salgsobjektet.
    rent_monthly : float
        Estimert månedlig leie (fra naboanalyse eller fallback).
    params : CashflowParams
        Investerings- og kostnadsparametere fra Streamlit-appen.
    template_path : str
        Sti til Boligkalkulator1.xlsx.
    tv_net_monthly : float
        TV og nett per måned (NOK).

    Returnerer
    ----------
    bytes
        Innholdet i den utfylte Excel-filen.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("openpyxl er ikke installert: pip install openpyxl") from e

    # -----------------------------------------------------------------------
    # Last inn template i en temporær kopi (bevarer all formatering)
    # -----------------------------------------------------------------------
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    shutil.copy2(template_path, tmp_path)
    wb = load_workbook(tmp_path)
    ws = wb["Bolig"]

    # -----------------------------------------------------------------------
    # Beregn verdier
    # -----------------------------------------------------------------------
    asking = _safe(estate.asking_price)
    total  = _safe(estate.total_price)
    # Bruk prisantydning som kjøpesum i templaten (slik som brukeren ville lagt inn)
    kjøpesum = asking if asking > 0 else total

    reg_charge = _safe(estate.registration_charge)
    # Dokumentavgift som andel – bare aktuelt for selveierboliger (~2.5%).
    # Hvis omkostninger er tilgjengelig bruker vi det direkte som transaksjonskostnader.
    if reg_charge > 0 and kjøpesum > 0:
        dok_avg_fraction = 0.0          # ikke dobbeltregn
        trans_cost = reg_charge
    else:
        dok_avg_fraction = 0.0
        trans_cost = 0.0

    equity    = _safe(params.equity)
    interest  = _safe(params.interest_rate, 0.055)
    insurance = _safe(params.insurance, 0.0)    # NOK/år → template forventer mnd-beløp i G9
    tax_year  = _safe(params.tax, 0.0)
    common_mnd = _safe(estate.common_monthly_cost, 0.0)
    maint_frac = _safe(params.maintenance_pct_rent, 0.05)

    # Strøm per måned
    area = _safe(estate.area, 0.0)
    if not params.tenant_pays_electricity and area > 0:
        elec_monthly = params.kwh_per_m2 * area * params.electricity_price / 12
    else:
        elec_monthly = 0.0

    energy_label = estate.energy_label or ""
    address = estate.location or estate.title or estate.finnkode

    # -----------------------------------------------------------------------
    # Skriv inn input-verdier (hardkodede celler – ikke formler)
    # -----------------------------------------------------------------------

    # Topp-info
    ws["K2"] = address

    # Kjøp
    ws["C2"] = kjøpesum
    ws["C4"] = dok_avg_fraction       # Dokumentavgift (andel)
    ws["C5"] = trans_cost             # Transaksjonskostnader
    ws["G5"] = equity                 # Egenkapital
    ws["G7"] = 0                      # Oppstartskostnad
    ws["G8"] = 0                      # Ombyggingskostnad

    # Rente
    ws["C7"] = interest

    # Driftskostnader
    ws["G9"]  = insurance / 12 if insurance > 0 else 0   # Forsikring pr. mnd (template trekker kun i jan)
    ws["G10"] = tax_year                                   # Eiendomsskatt (år)
    ws["G11"] = _safe(estate.municipality_cost_year)       # Kommunale avgifter (år)
    ws["G12"] = common_mnd                                 # Felleskost pr. mnd

    # Strøm og energi
    ws["K6"] = tv_net_monthly          # TV og nett
    ws["K8"] = round(elec_monthly, 0)  # Strøm pr. mnd
    ws["K9"] = energy_label            # Energiklasse

    # Leie og vedlikehold
    ws["C13"] = round(rent_monthly, 0)  # Månedsleie
    ws["C15"] = maint_frac              # Vedlikeholdskostnad (andel av leie)

    # -----------------------------------------------------------------------
    # Lagre til bytes-buffer
    # -----------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Rydd opp temp-fil
    Path(tmp_path).unlink(missing_ok=True)

    return buf.getvalue()
